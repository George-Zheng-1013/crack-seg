"""
报告生成模块
支持：PDF 报告（含缩略图）、Excel 报告
"""

import io
import json
import datetime
from pathlib import Path
from typing import Optional

from app.inference import FrameResult, Detection


API_CONFIG_PATH = Path(__file__).resolve().parent.parent / "api.json"


def _feature(det: Detection, key: str, default=""):
    return (det.features or {}).get(key, default)


def _top_match(det: Detection) -> dict:
    return (det.cause_analysis or {}).get("top_match", {}) or {}


def _top_feature_text(det: Detection) -> str:
    top = _top_match(det)
    return top.get("feature_cn") or top.get("label_cn", "")


def _cause_text(det: Detection, key: str, sep: str = "、") -> str:
    values = (det.cause_analysis or {}).get(key, [])
    return sep.join(values) if values else ""


def _pdf_advice_id(seq: int, frame_result: FrameResult, det: Detection) -> str:
    return f"{seq}:{frame_result.frame_index}:{det.det_id}"


def _display_frame_index(frame_result: FrameResult, det: Detection) -> int:
    if det.best_frame_index is not None:
        return int(det.best_frame_index)
    if det.first_frame is not None:
        return int(det.first_frame)
    return int(frame_result.frame_index)


def _load_deepseek_config() -> dict:
    try:
        config = json.loads(API_CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}

    deepseek = config.get("deepseek", {}) if isinstance(config, dict) else {}
    api_key = str(deepseek.get("api_key", "")).strip()
    if not api_key or api_key == "<your api key>":
        return {}

    return {
        "api_key": api_key,
        "base_url": deepseek.get("base_url", "https://api.deepseek.com"),
        "model": deepseek.get("model", "deepseek-v4-pro"),
        "timeout_sec": deepseek.get("timeout_sec", 30),
    }


def _llm_request_options() -> dict:
    return {
        "extra_body": {
            "thinking": {"type": "disabled"},
        },
    }


def _pdf_advice_payload(results: list[FrameResult]) -> list[dict]:
    items = []
    seq = 1
    for frame_result in results:
        for det in frame_result.detections:
            top = _top_match(det)
            items.append({
                "id": _pdf_advice_id(seq, frame_result, det),
                "类别": det.class_name,
                "置信度": round(det.confidence, 4),
                "掩码面积": det.mask_area_px,
                "视觉特征": _top_feature_text(det),
                "匹配得分": top.get("score", ""),
                "可能成因": _cause_text(det, "possible_causes"),
                "原始排查建议": _cause_text(det, "inspection_advice", sep="；"),
                "面积比": _feature(det, "rectangularity"),
                "细长度": _feature(det, "slenderness"),
                "主方向": _feature(det, "major_direction"),
                "边界复杂度": _feature(det, "boundary_complexity"),
                "骨架长度": _feature(det, "skeleton_length"),
                "分支数": _feature(det, "branch_points"),
                "端点数": _feature(det, "end_points"),
            })
            seq += 1
    return items


def _generate_pdf_api_advice(results: list[FrameResult]) -> dict[str, str]:
    config = _load_deepseek_config()
    if not config:
        return {}

    items = _pdf_advice_payload(results)
    if not items:
        return {}

    system_prompt = """
你是混凝土和基础设施表观缺陷排查助手。请根据用户提供的每个检测实例信息，综合类别、置信度、掩码面积、视觉特征、可能成因、原始排查建议以及形态指标，生成更具体的中文排查建议。

要求：
1. 只输出 JSON，不输出解释性文本。
2. JSON 格式必须为 {"items":[{"id":"...","inspection_advice":"..."}]}。
3. 每条 inspection_advice 控制在 1-2 句，面向现场复核和工程排查。
4. 不把模型匹配结果写成确定诊断，使用“建议复核”“重点排查”“结合记录核查”等措辞。
5. 不要只复述原始排查建议，要综合形态指标、视觉特征和可能成因。
"""
    user_prompt = json.dumps({"items": items}, ensure_ascii=False)

    try:
        from openai import OpenAI

        client = OpenAI(
            api_key=config["api_key"],
            base_url=config["base_url"],
            timeout=float(config["timeout_sec"]),
        )
        response = client.chat.completions.create(
            model=config["model"],
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            response_format={"type": "json_object"},
            **_llm_request_options(),
        )
        content = response.choices[0].message.content or "{}"
        parsed = json.loads(content)
        api_items = parsed.get("items", [])
        if not isinstance(api_items, list):
            return {}

        advice_by_id = {}
        valid_ids = {item["id"] for item in items}
        for item in api_items:
            if not isinstance(item, dict):
                continue
            item_id = str(item.get("id", ""))
            advice = str(item.get("inspection_advice", "")).strip()
            if item_id in valid_ids and advice:
                advice_by_id[item_id] = advice
        return advice_by_id
    except Exception:
        return {}


def _scene_payload(results: list[FrameResult], video_summary: Optional[dict]) -> dict:
    items = []
    seq = 1
    for frame_result in results:
        for det in frame_result.detections:
            top = _top_match(det)
            items.append({
                "id": str(seq),
                "track_id": det.track_id,
                "类别": det.class_name,
                "置信度": round(det.confidence, 4),
                "代表帧": _display_frame_index(frame_result, det),
                "首次出现时间": det.first_time,
                "末次出现时间": det.last_time,
                "出现帧数": det.frame_count,
                "掩码面积": det.mask_area_px,
                "视觉特征": _top_feature_text(det),
                "匹配得分": top.get("score", ""),
                "可能成因": _cause_text(det, "possible_causes"),
                "排查建议": _cause_text(det, "inspection_advice", sep="；"),
                "面积比": _feature(det, "rectangularity"),
                "细长度": _feature(det, "slenderness"),
                "主方向": _feature(det, "major_direction"),
                "边界复杂度": _feature(det, "boundary_complexity"),
                "骨架长度": _feature(det, "skeleton_length"),
                "分支数": _feature(det, "branch_points"),
                "端点数": _feature(det, "end_points"),
            })
            seq += 1
    return {
        "video_summary": video_summary or {},
        "items": items,
    }


def _fallback_scene_analysis(results: list[FrameResult], video_summary: Optional[dict]) -> str:
    all_dets = [d for r in results for d in r.detections]
    if not all_dets:
        return "该视频场景未检测到明确缺陷实例，建议结合原始视频画面、拍摄范围和现场巡检记录继续复核。"
    by_class: dict[str, int] = {}
    for det in all_dets:
        by_class[det.class_name] = by_class.get(det.class_name, 0) + 1
    class_text = "、".join(f"{k}{v}处" for k, v in sorted(by_class.items(), key=lambda x: -x[1]))
    sampled = (video_summary or {}).get("sampled_frames", "")
    return (
        f"该视频场景共汇总到 {len(all_dets)} 个唯一缺陷实例，类别分布为 {class_text}。"
        f"建议结合视频连续画面、代表帧 crop、缺陷位置关系和施工/渗排水记录进行复核；"
        f"本结论基于 {sampled or '若干'} 个采样跟踪帧形成，作为现场排查的辅助依据。"
    )


def _generate_video_scene_analysis(results: list[FrameResult], video_summary: Optional[dict]) -> str:
    if not any(d.track_id is not None for r in results for d in r.detections):
        return ""

    config = _load_deepseek_config()
    payload = _scene_payload(results, video_summary)
    if not config or not payload["items"]:
        return _fallback_scene_analysis(results, video_summary)

    system_prompt = """
你是混凝土和基础设施表观缺陷排查助手。用户会提供同一视频场景中跟踪得到的唯一缺陷实例、类别分布、时间范围、形态指标、视觉特征匹配、可能成因和原始建议。
请输出 JSON：{"scene_analysis":"..."}。
要求：1. 使用中文；2. 写成一段综合分析，约 120-220 字；3. 强调这是同一视频场景下的辅助研判，不写成确定诊断；4. 综合类别、数量、时间范围、形态指标和可能成因，不只复述单条建议；5. 给出现场复核重点。
"""
    try:
        from openai import OpenAI

        client = OpenAI(
            api_key=config["api_key"],
            base_url=config["base_url"],
            timeout=float(config["timeout_sec"]),
        )
        response = client.chat.completions.create(
            model=config["model"],
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": json.dumps(payload, ensure_ascii=False)},
            ],
            response_format={"type": "json_object"},
            **_llm_request_options(),
        )
        parsed = json.loads(response.choices[0].message.content or "{}")
        text = str(parsed.get("scene_analysis", "")).strip()
        return text or _fallback_scene_analysis(results, video_summary)
    except Exception:
        return _fallback_scene_analysis(results, video_summary)


# ──────────────────────────────────────────────
# Excel 报告
# ──────────────────────────────────────────────

def generate_excel_report(
    results: list[FrameResult],
    source_name: str = "unknown",
    output_path: Optional[str] = None,
) -> bytes:
    """
    生成 Excel 检测报告。

    Sheet1 - 汇总：总缺陷数、各类别统计、推理时间
    Sheet2 - 明细：每个缺陷的详细信息（类型、坐标、置信度等）

    :param results:      FrameResult 列表
    :param source_name:  来源文件名（显示在报告标题）
    :param output_path:  若提供则同时写入文件
    :return: Excel 文件 bytes
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, Alignment, PatternFill, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()

    # ── 配色方案 ──
    COLOR_HEADER_BG   = "1F3864"   # 深蓝
    COLOR_HEADER_FONT = "FFFFFF"   # 白
    COLOR_SUBHEADER   = "2E75B6"   # 中蓝
    COLOR_ALT_ROW     = "EBF3FB"   # 浅蓝（交替行）
    COLOR_WARN        = "FF0000"   # 红色（高风险）
    COLOR_OK          = "00B050"   # 绿色

    def header_style(cell, bg=COLOR_HEADER_BG):
        cell.font      = Font(bold=True, color=COLOR_HEADER_FONT, size=11)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="AAAAAA")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def normal_style(cell, bg="FFFFFF"):
        cell.alignment = Alignment(horizontal="left", vertical="center")
        thin = Side(style="thin", color="DDDDDD")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.fill = PatternFill("solid", fgColor=bg)

    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ═══════════════════════════════════════════
    # Sheet 1：检测汇总
    # ═══════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = "检测汇总"
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 22

    row = 1
    # 大标题
    ws_sum.merge_cells(f"A{row}:B{row}")
    c = ws_sum[f"A{row}"]
    c.value = "基础设施外观缺陷检测报告"
    c.font = Font(bold=True, size=15, color=COLOR_SUBHEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[row].height = 30
    row += 1

    # 元信息
    meta = [
        ("报告生成时间", now_str),
        ("检测来源",     source_name),
        ("处理帧/图数",  str(len(results))),
    ]
    total_defects = sum(r.total_defects for r in results)
    avg_conf = 0.0
    all_dets = [d for r in results for d in r.detections]
    if all_dets:
        avg_conf = sum(d.confidence for d in all_dets) / len(all_dets)
    avg_infer = sum(r.inference_time_ms for r in results) / max(len(results), 1)

    meta += [
        ("总检测缺陷数",  str(total_defects)),
        ("平均置信度",    f"{avg_conf:.3f}"),
        ("平均推理时间",  f"{avg_infer:.1f} ms"),
    ]

    for k, v in meta:
        ws_sum[f"A{row}"] = k
        ws_sum[f"B{row}"] = v
        normal_style(ws_sum[f"A{row}"])
        normal_style(ws_sum[f"B{row}"])
        row += 1

    row += 1

    # 各类别统计
    ws_sum.merge_cells(f"A{row}:B{row}")
    c = ws_sum[f"A{row}"]
    c.value = "各类别缺陷统计"
    header_style(c, bg=COLOR_SUBHEADER)
    row += 1

    by_class: dict[str, int] = {}
    for r in results:
        for cls, cnt in r.by_class.items():
            by_class[cls] = by_class.get(cls, 0) + cnt

    ws_sum[f"A{row}"] = "类别名称"
    ws_sum[f"B{row}"] = "检测数量"
    header_style(ws_sum[f"A{row}"])
    header_style(ws_sum[f"B{row}"])
    row += 1

    for cls_name, cnt in sorted(by_class.items(), key=lambda x: -x[1]):
        ws_sum[f"A{row}"] = cls_name
        ws_sum[f"B{row}"] = cnt
        normal_style(ws_sum[f"A{row}"])
        ws_sum[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws_sum[f"B{row}"].fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)
        row += 1

    # ═══════════════════════════════════════════
    # Sheet 2：缺陷明细
    # ═══════════════════════════════════════════
    ws_det = wb.create_sheet("缺陷明细")

    col_headers = [
        ("序号",       8),
        ("帧/图编号",  12),
        ("时间戳(s)",  12),
        ("类别",       14),
        ("置信度",     10),
        ("X1(px)",     10),
        ("Y1(px)",     10),
        ("X2(px)",     10),
        ("Y2(px)",     10),
        ("中心X(px)",  10),
        ("中心Y(px)",  10),
        ("宽度(px)",   10),
        ("高度(px)",   10),
        ("边界框面积", 14),
        ("掩码面积",   12),
        ("面积比",     10),
        ("细长度",     10),
        ("主方向(°)",  12),
        ("边界复杂度", 12),
        ("骨架长度",   12),
        ("分支数",     10),
        ("端点数",     10),
        ("视觉特征",   22),
        ("匹配得分",   10),
        ("可能成因",   32),
        ("排查建议",   40),
        ("推理时间(ms)", 14),
    ]

    for col_i, (header, width) in enumerate(col_headers, start=1):
        cell = ws_det.cell(row=1, column=col_i, value=header)
        header_style(cell)
        ws_det.column_dimensions[get_column_letter(col_i)].width = width
    ws_det.row_dimensions[1].height = 22

    det_row = 2
    global_idx = 1
    for frame_result in results:
        for det in frame_result.detections:
            top = _top_match(det)
            bg = COLOR_ALT_ROW if det_row % 2 == 0 else "FFFFFF"
            row_data = [
                global_idx,
                frame_result.frame_index,
                frame_result.timestamp_sec,
                det.class_name,
                round(det.confidence, 4),
                det.bbox[0],
                det.bbox[1],
                det.bbox[2],
                det.bbox[3],
                det.center[0],
                det.center[1],
                det.width,
                det.height,
                det.area_px,
                det.mask_area_px,
                _feature(det, "rectangularity"),
                _feature(det, "slenderness"),
                _feature(det, "major_direction"),
                _feature(det, "boundary_complexity"),
                _feature(det, "skeleton_length"),
                _feature(det, "branch_points"),
                _feature(det, "end_points"),
                _top_feature_text(det),
                top.get("score", ""),
                _cause_text(det, "possible_causes"),
                _cause_text(det, "inspection_advice", sep="；"),
                round(frame_result.inference_time_ms, 2),
            ]
            for col_i, val in enumerate(row_data, start=1):
                cell = ws_det.cell(row=det_row, column=col_i, value=val)
                normal_style(cell, bg=bg)
                if col_i == 5 and val < 0.5:          # 低置信度标红
                    cell.font = Font(color=COLOR_WARN)
                elif col_i == 5 and val >= 0.8:        # 高置信度标绿
                    cell.font = Font(color=COLOR_OK, bold=True)

            det_row += 1
            global_idx += 1

    # ── 冻结首行 ──
    ws_det.freeze_panes = "A2"

    # ── 输出 ──
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    if output_path:
        Path(output_path).write_bytes(raw)

    return raw


# ──────────────────────────────────────────────
# PDF 报告
# ──────────────────────────────────────────────

def generate_pdf_report(
    results: list[FrameResult],
    source_name: str = "unknown",
    annotated_images: Optional[list] = None,   # list of BGR numpy arrays
    instance_image_paths: Optional[list[dict]] = None,
    video_summary: Optional[dict] = None,
    output_path: Optional[str] = None,
    max_images: int = 10,
) -> bytes:
    """
    生成 PDF 检测报告。

    包含：封面、统计摘要、检测明细表、（可选）标注图像

    :param results:           FrameResult 列表
    :param source_name:       来源名称
    :param annotated_images:  标注后的图像列表（BGR numpy array），用于嵌入 PDF
    :param output_path:       若提供则写入文件
    :param max_images:        PDF 中最多嵌入几张图
    :return: PDF bytes
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable, Image as RLImage, PageBreak,
        )
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        raise ImportError("请安装 reportlab: pip install reportlab")

    import cv2
    import tempfile
    import os

    # ── 注册中文字体（自动检测系统字体）──
    _register_cjk_font()

    from reportlab.lib.styles import getSampleStyleSheet
    styles = getSampleStyleSheet()

    FONT_CN = _get_cjk_font_name()

    def cn_style(size=10, bold=False, color=colors.black, align="left"):
        align_map = {"left": 0, "center": 1, "right": 2}
        return ParagraphStyle(
            "cn",
            fontName=FONT_CN if not bold else FONT_CN,
            fontSize=size,
            textColor=color,
            alignment=align_map.get(align, 0),
            leading=size * 1.4,
        )

    # ── 汇总数据 ──
    all_dets   = [d for r in results for d in r.detections]
    total_def  = len(all_dets)
    avg_conf   = sum(d.confidence for d in all_dets) / max(len(all_dets), 1)
    avg_infer  = sum(r.inference_time_ms for r in results) / max(len(results), 1)
    by_class: dict[str, int] = {}
    for r in results:
        for cls, cnt in r.by_class.items():
            by_class[cls] = by_class.get(cls, 0) + cnt
    now_str = datetime.datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")

    buf   = io.BytesIO()
    doc   = SimpleDocTemplate(
        buf,
        pagesize      = A4,
        leftMargin    = 20 * mm,
        rightMargin   = 20 * mm,
        topMargin     = 20 * mm,
        bottomMargin  = 20 * mm,
    )
    story = []
    W_pt  = A4[0] - 40 * mm   # 可用宽度
    H_pt  = A4[1] - 40 * mm   # 可用高度

    def fit_image_size(img_w: int, img_h: int, max_w: float, max_h: float) -> tuple[float, float]:
        if img_w <= 0 or img_h <= 0:
            return max_w, max_h
        scale = min(max_w / img_w, max_h / img_h, 1.0)
        return img_w * scale, img_h * scale

    # ─────────── 封面 ───────────
    story.append(Spacer(1, 30 * mm))
    story.append(Paragraph("基础设施外观缺陷", cn_style(28, bold=True, color=colors.HexColor("#1F3864"), align="center")))
    story.append(Paragraph("智能检测报告", cn_style(28, bold=True, color=colors.HexColor("#1F3864"), align="center")))
    story.append(Spacer(1, 8 * mm))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#2E75B6")))
    story.append(Spacer(1, 6 * mm))

    scene_analysis = _generate_video_scene_analysis(results, video_summary)
    if scene_analysis:
        story.append(Paragraph("视频场景综合分析", cn_style(12, bold=True, color=colors.HexColor("#2E75B6"))))
        story.append(Spacer(1, 2 * mm))
        story.append(Paragraph(scene_analysis, cn_style(9)))
        story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(f"检测来源：{source_name}", cn_style(12, align="center")))
    story.append(Paragraph(f"生成时间：{now_str}", cn_style(12, align="center")))
    story.append(PageBreak())

    # ─────────── 统计摘要 ───────────
    story.append(Paragraph("1. 检测统计摘要", cn_style(14, bold=True, color=colors.HexColor("#2E75B6"))))
    story.append(Spacer(1, 4 * mm))

    summary_data = [
        ["指标", "数值"],
        ["处理帧/图数",  str(len(results))],
        ["总检测缺陷数", str(total_def)],
        ["平均置信度",   f"{avg_conf:.3f}"],
        ["平均推理时间", f"{avg_infer:.1f} ms"],
    ]
    for cls_name, cnt in sorted(by_class.items(), key=lambda x: -x[1]):
        summary_data.append([f"  {cls_name} 数量", str(cnt)])

    ts = TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#1F3864")),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",    (0, 0), (-1, -1), FONT_CN),
        ("FONTSIZE",    (0, 0), (-1, 0),  11),
        ("FONTSIZE",    (0, 1), (-1, -1), 10),
        ("ALIGN",       (0, 0), (-1, -1), "LEFT"),
        ("ALIGN",       (1, 0), (1, -1),  "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF3FB")]),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#AAAAAA")),
        ("TOPPADDING",  (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ])
    t = Table(summary_data, colWidths=[W_pt * 0.55, W_pt * 0.45])
    t.setStyle(ts)
    story.append(t)
    story.append(Spacer(1, 6 * mm))

    # ─────────── 明细表 ───────────
    story.append(Paragraph("2. 缺陷明细列表", cn_style(14, bold=True, color=colors.HexColor("#2E75B6"))))
    story.append(Spacer(1, 4 * mm))

    col_w = [W_pt * x for x in [0.05, 0.09, 0.11, 0.08, 0.11, 0.13, 0.19, 0.24]]
    det_headers = ["#", "帧编号", "类别", "置信度", "掩码面积", "视觉特征", "可能成因", "排查建议"]
    det_data = [det_headers]

    api_advice = _generate_pdf_api_advice(results)
    det_seq = 1
    for r in results:
        for d in r.detections:
            top = _top_match(d)
            advice_id = _pdf_advice_id(det_seq, r, d)
            advice_text = api_advice.get(
                advice_id,
                _cause_text(d, "inspection_advice", sep="；"),
            )
            det_data.append([
                str(d.det_id),
                str(_display_frame_index(r, d)),
                d.class_name,
                f"{d.confidence:.3f}",
                str(d.mask_area_px),
                f"{_top_feature_text(d)} {top.get('score', '')}",
                _cause_text(d, "possible_causes"),
                advice_text,
            ])
            det_seq += 1

    det_data = [
        [
            Paragraph(str(cell), cn_style(8 if row_idx == 0 else 7, align="center"))
            for cell in row
        ]
        for row_idx, row in enumerate(det_data)
    ]

    det_ts = TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#2E75B6")),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, -1), FONT_CN),
        ("FONTSIZE",      (0, 0), (-1, -1), 7),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF3FB")]),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ])
    t2 = Table(det_data, colWidths=col_w, repeatRows=1)
    t2.setStyle(det_ts)
    story.append(t2)

    if all_dets:
        story.append(Spacer(1, 5 * mm))
        story.append(Paragraph("3. 裂缝形态特征", cn_style(14, bold=True, color=colors.HexColor("#2E75B6"))))
        story.append(Spacer(1, 4 * mm))

        feat_data = [["#", "面积比", "细长度", "主方向", "边界复杂度", "骨架长度", "分支数", "端点数"]]
        for r in results:
            for d in r.detections:
                feat_data.append([
                    str(d.det_id),
                    str(_feature(d, "rectangularity")),
                    str(_feature(d, "slenderness")),
                    str(_feature(d, "major_direction")),
                    str(_feature(d, "boundary_complexity")),
                    str(_feature(d, "skeleton_length")),
                    str(_feature(d, "branch_points")),
                    str(_feature(d, "end_points")),
                ])
        feat_data = [
            [Paragraph(str(cell), cn_style(7, align="center")) for cell in row]
            for row in feat_data
        ]
        feat_table = Table(feat_data, colWidths=[W_pt / 8] * 8, repeatRows=1)
        feat_table.setStyle(det_ts)
        story.append(feat_table)

    # ─────────── 标注图像 ───────────
    if instance_image_paths:
        story.append(PageBreak())
        story.append(Paragraph("4. 缺陷实例局部图像", cn_style(14, bold=True, color=colors.HexColor("#2E75B6"))))
        story.append(Spacer(1, 4 * mm))
        for item in instance_image_paths[:max_images]:
            img_path = Path(str(item.get("path", "")))
            if not img_path.exists() or not img_path.is_file():
                continue
            try:
                import cv2 as _cv2
                img_bgr = _cv2.imread(str(img_path))
                if img_bgr is None:
                    continue
                H_img, W_img = img_bgr.shape[:2]
                disp_w, disp_h = fit_image_size(W_img, H_img, W_pt, H_pt * 0.72)
                story.append(RLImage(str(img_path), width=disp_w, height=disp_h))
                if item.get("track_id") is not None:
                    caption = (
                        f"实例 {item.get('det_id', '')} | track T{item.get('track_id', '-')} | "
                        f"{item.get('class_name', '')} | 代表帧 {item.get('frame_index', '-')} | "
                        f"{item.get('first_time', '-')}s - {item.get('last_time', '-')}s"
                    )
                else:
                    caption = (
                        f"实例 {item.get('det_id', '')} | {item.get('class_name', '')} | "
                        f"图像序号 {item.get('frame_index', '-')}"
                    )
                story.append(Paragraph(caption, cn_style(8, align="center")))
                story.append(Spacer(1, 4 * mm))
            except Exception:
                continue

    if annotated_images:
        story.append(PageBreak())
        story.append(Paragraph("4. 检测标注图像", cn_style(14, bold=True, color=colors.HexColor("#2E75B6"))))
        story.append(Spacer(1, 4 * mm))

        tmp_files = []
        try:
            for i, (img_bgr, frame_res) in enumerate(
                zip(annotated_images[:max_images], results[:max_images])
            ):
                import cv2 as _cv2
                H_img, W_img = img_bgr.shape[:2]
                disp_w, disp_h = fit_image_size(W_img, H_img, W_pt, H_pt * 0.72)

                tmpf = tempfile.NamedTemporaryFile(suffix=".jpg", delete=False)
                tmp_files.append(tmpf.name)
                _cv2.imwrite(tmpf.name, img_bgr, [_cv2.IMWRITE_JPEG_QUALITY, 100])
                tmpf.close()

                caption = (f"图 {i+1} | 帧编号: {frame_res.frame_index} | "
                           f"检测缺陷: {frame_res.total_defects} 处 | "
                           f"推理时间: {frame_res.inference_time_ms:.1f} ms")
                story.append(RLImage(tmpf.name, width=disp_w, height=disp_h))
                story.append(Paragraph(caption, cn_style(8, align="center")))
                story.append(Spacer(1, 4 * mm))
        finally:
            pass  # 延迟删除，build 之后再清理

        doc.build(story)
        raw = buf.getvalue()

        for f in tmp_files:
            try:
                os.unlink(f)
            except Exception:
                pass
    else:
        doc.build(story)
        raw = buf.getvalue()

    if output_path:
        Path(output_path).write_bytes(raw)

    return raw


# ──────────────────────────────────────────────
# 字体辅助
# ──────────────────────────────────────────────

_CJK_FONT_REGISTERED = False
_CJK_FONT_NAME = "Helvetica"   # fallback


def _register_cjk_font():
    """尝试注册系统 CJK 字体，供 ReportLab 使用"""
    global _CJK_FONT_REGISTERED, _CJK_FONT_NAME
    if _CJK_FONT_REGISTERED:
        return

    import sys
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    candidates = []
    if sys.platform == "win32":
        import winreg, os
        candidates = [
            r"C:\Windows\Fonts\msyh.ttc",       # 微软雅黑
            r"C:\Windows\Fonts\simhei.ttf",     # 黑体
            r"C:\Windows\Fonts\simsun.ttc",     # 宋体
            r"C:\Windows\Fonts\simfang.ttf",    # 仿宋
        ]
    elif sys.platform == "darwin":
        candidates = [
            "/Library/Fonts/Arial Unicode.ttf",
            "/System/Library/Fonts/PingFang.ttc",
        ]
    else:  # Linux
        candidates = [
            "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
            "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJKsc-Regular.otf",
        ]

    for path in candidates:
        try:
            pdfmetrics.registerFont(TTFont("CJKFont", path))
            _CJK_FONT_NAME = "CJKFont"
            _CJK_FONT_REGISTERED = True
            return
        except Exception:
            continue

    _CJK_FONT_REGISTERED = True   # 使用 fallback


def _get_cjk_font_name() -> str:
    _register_cjk_font()
    return _CJK_FONT_NAME
